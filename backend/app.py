import os
import uuid
import sqlite3
import jwt
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
import piexif
from werkzeug.security import generate_password_hash, check_password_hash

# Document libraries
from pypdf import PdfReader, PdfWriter
from docx import Document
from openpyxl import load_workbook
from pptx import Presentation

app = Flask(__name__)
CORS(app)

SECRET_KEY = "metaguard_secret_key" # In production, use environment variable

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
CLEAN_FOLDER = os.path.join(os.path.dirname(__file__), 'cleaned')
DB_PATH = os.path.join(os.path.dirname(__file__), 'database.db')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CLEAN_FOLDER, exist_ok=True)

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE,
            password_hash TEXT,
            created_at TEXT
        )
    ''')
    try:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
    except sqlite3.OperationalError:
        pass # Column might already exist
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS history (
            id TEXT PRIMARY KEY,
            user_id TEXT,
            filename TEXT,
            upload_date TEXT,
            score INTEGER,
            risk_level TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    ''')
    # Migrate history to add user_id if not exists
    try:
        c.execute("ALTER TABLE history ADD COLUMN user_id TEXT")
    except sqlite3.OperationalError:
        pass # Column might already exist
    conn.commit()
    conn.close()

init_db()

def get_optional_user():
    token = request.headers.get('Authorization')
    if token and token.startswith('Bearer '):
        token = token.split(" ")[1]
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            return data['user_id']
        except:
            return None
    return None

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_optional_user()
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(user_id, *args, **kwargs)
    return decorated

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    username, email, password = data.get('username'), data.get('email'), data.get('password')
    if not username or not email or not password:
         return jsonify({'error': 'Missing required fields'}), 400
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        user_id = str(uuid.uuid4())
        pw_hash = generate_password_hash(password)
        c.execute("INSERT INTO users (id, username, email, password_hash, created_at) VALUES (?, ?, ?, ?, ?)",
                  (user_id, username, email, pw_hash, datetime.now(timezone.utc).isoformat()))
        conn.commit()
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username or email already exists'}), 400
    finally:
        conn.close()
        
    return jsonify({'message': 'User registered successfully'})

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username, password = data.get('username'), data.get('password')
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, password_hash FROM users WHERE username = ?", (username,))
    user = c.fetchone()
    conn.close()
    
    if not user or not check_password_hash(user[1], password):
         return jsonify({'error': 'Invalid credentials'}), 401
         
    token = jwt.encode({
        'user_id': user[0],
        'exp': datetime.now(timezone.utc) + timedelta(days=7)
    }, SECRET_KEY, algorithm="HS256")
    
    return jsonify({'token': token, 'username': username})

def get_geotagging(exif):
    if not exif:
        return None
    geotagging = {}
    for (idx, tag) in TAGS.items():
        if tag == 'GPSInfo':
            if idx not in exif:
                return None
            for (key, val) in GPSTAGS.items():
                if key in exif[idx]:
                    geotagging[val] = exif[idx][key]
    return geotagging

def extract_metadata_image(image_path, filename):
    try:
        image = Image.open(image_path)
        exif_raw = image._getexif()
        metadata = []
        high_risk, med_risk, low_risk = 0, 0, 0
        
        metadata.append({"key": "Format", "value": image.format, "risk": "Low"})
        metadata.append({"key": "Size", "value": f"{os.path.getsize(image_path) / 1024:.2f} KB", "risk": "Low"})
        metadata.append({"key": "Dimensions", "value": f"{image.width}x{image.height}", "risk": "Low"})
        low_risk += 3

        if not exif_raw:
            return metadata, 100, "Safe"

        for tag_id in exif_raw:
            tag = TAGS.get(tag_id, tag_id)
            data = exif_raw.get(tag_id)
            if isinstance(data, bytes):
                try:
                    data = data.decode()
                except:
                    data = "<binary data>"
            
            risk = "Low"
            if tag in ['Model', 'Make', 'BodySerialNumber', 'LensMake', 'LensModel']:
                risk = "High"
                high_risk += 1
            elif 'Date' in str(tag) or 'Time' in str(tag) or tag == 'Software':
                risk = "Medium"
                med_risk += 1
            elif tag != 'GPSInfo':
                low_risk += 1
                
            if tag != 'GPSInfo' and len(str(data)) < 100:
                metadata.append({"key": tag, "value": str(data), "risk": risk})
                
        geo_info = get_geotagging(exif_raw)
        if geo_info:
            metadata.append({"key": "GPS Coordinates", "value": "Present", "risk": "Critical"})
            high_risk += 2
            for g_key, g_val in geo_info.items():
                 metadata.append({"key": f"GPS {g_key}", "value": str(g_val), "risk": "Critical"})
                 
        score = max(0, 100 - (high_risk * 25) - (med_risk * 10))
        risk_level = "High" if high_risk > 0 else "Medium" if med_risk > 0 else "Low"
        return metadata, score, risk_level
    except Exception as e:
        print(f"Error extracting image metadata: {e}")
        return [{"key": "Error", "value": "Could not extract metadata", "risk": "Low"}], 100, "Safe"

        
def extract_metadata_doc(file_path, filename, ext):
    metadata = []
    high_risk, med_risk, low_risk = 0, 0, 0
    metadata.append({"key": "Format", "value": ext.upper(), "risk": "Low"})
    metadata.append({"key": "Size", "value": f"{os.path.getsize(file_path) / 1024:.2f} KB", "risk": "Low"})
    low_risk += 2
    
    try:
        if ext == 'pdf':
            reader = PdfReader(file_path)
            meta = reader.metadata
            if meta:
                for k, v in meta.items():
                    k_str = str(k).strip('/')
                    v_str = str(v)
                    risk = "Medium"
                    if k_str in ['Author', 'Creator', 'Producer']: risk = "High"; high_risk += 1
                    else: med_risk += 1
                    metadata.append({"key": k_str, "value": v_str[:100], "risk": risk})
        elif ext == 'docx':
            doc = Document(file_path)
            props = doc.core_properties
            if props.author: metadata.append({"key": "Author", "value": props.author, "risk": "High"}); high_risk += 1
            if props.last_modified_by: metadata.append({"key": "Last Modified By", "value": props.last_modified_by, "risk": "High"}); high_risk += 1
            if props.created: metadata.append({"key": "Created", "value": str(props.created), "risk": "Medium"}); med_risk += 1
            if props.modified: metadata.append({"key": "Modified", "value": str(props.modified), "risk": "Medium"}); med_risk += 1
        elif ext == 'xlsx':
            wb = load_workbook(file_path)
            props = wb.properties
            if props.creator: metadata.append({"key": "Creator", "value": props.creator, "risk": "High"}); high_risk += 1
            if props.lastModifiedBy: metadata.append({"key": "Last Modified By", "value": props.lastModifiedBy, "risk": "High"}); high_risk += 1
            if props.created: metadata.append({"key": "Created", "value": str(props.created), "risk": "Medium"}); med_risk += 1
        elif ext == 'pptx':
            prs = Presentation(file_path)
            props = prs.core_properties
            if props.author: metadata.append({"key": "Author", "value": props.author, "risk": "High"}); high_risk += 1
            if props.last_modified_by: metadata.append({"key": "Last Modified By", "value": props.last_modified_by, "risk": "High"}); high_risk += 1
            if props.created: metadata.append({"key": "Created", "value": str(props.created), "risk": "Medium"}); med_risk += 1
    except Exception as e:
        print(f"Error extracting doc metadata: {e}")
        
    score = max(0, 100 - (high_risk * 25) - (med_risk * 10))
    risk_level = "High" if high_risk > 0 else "Medium" if med_risk > 0 else "Low"
    return metadata, score, risk_level

def clean_image(input_path, output_path, share_safe=False):
    image = Image.open(input_path)
    
    if share_safe:
        # Resize to max 1080p for social media sharing
        max_size = (1080, 1080)
        image.thumbnail(max_size, Image.Resampling.LANCZOS)
    
    data = list(image.getdata())
    image_without_exif = Image.new(image.mode, image.size)
    image_without_exif.putdata(data)
    
    if share_safe and image.format in ['JPEG', 'PNG']:
        # Compress and save as JPEG to remove any extra metadata embedded in PNG chunks
        # Convert RGBA to RGB if needed
        if image_without_exif.mode in ("RGBA", "P"):
            image_without_exif = image_without_exif.convert("RGB")
        image_without_exif.save(output_path, "JPEG", quality=80)
        return
        
    if image.format == 'JPEG':
        try:
             piexif.remove(input_path, output_path)
        except:
             image_without_exif.save(output_path, format=image.format)
    else:
        if image.format == 'PNG':
            image.save(output_path, "PNG", pnginfo=None)
        else:
            image_without_exif.save(output_path, format=image.format or 'JPEG')

def clean_doc(input_path, output_path, ext):
    try:
        if ext == 'pdf':
            reader = PdfReader(input_path)
            writer = PdfWriter()
            for page in reader.pages: writer.add_page(page)
            writer.add_metadata({}) # Clear metadata
            with open(output_path, "wb") as f:
                writer.write(f)
        elif ext == 'docx':
            doc = Document(input_path)
            doc.core_properties.author = ""
            doc.core_properties.last_modified_by = ""
            doc.core_properties.comments = ""
            doc.save(output_path)
        elif ext == 'xlsx':
            wb = load_workbook(input_path)
            wb.properties.creator = ""
            wb.properties.lastModifiedBy = ""
            wb.save(output_path)
        elif ext == 'pptx':
            prs = Presentation(input_path)
            prs.core_properties.author = ""
            prs.core_properties.last_modified_by = ""
            prs.save(output_path)
    except Exception as e:
        print(f"Doc cleaning error: {e}")
        # Falback copy
        import shutil
        shutil.copyfile(input_path, output_path)


@app.route('/api/analyze', methods=['POST'])
def analyze():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    share_safe = request.form.get('share_safe', 'false').lower() == 'true'
    
    ext = file.filename.split('.')[-1].lower()
    allowed_images = ['jpg', 'jpeg', 'png']
    allowed_docs = ['pdf', 'docx', 'xlsx', 'pptx']
    
    if ext not in allowed_images and ext not in allowed_docs:
        return jsonify({"error": "Unsupported file format"}), 400

    file_id = str(uuid.uuid4())
    filename = f"{file_id}.{ext}"
    input_path = os.path.join(UPLOAD_FOLDER, filename)
    
    # Enforce safe filename, but we are using GUID anyway
    output_filename = filename
    if share_safe and ext in allowed_images:
        output_filename = f"{file_id}.jpg" # share safe converts to jpg
        
    output_path = os.path.join(CLEAN_FOLDER, output_filename)
    file.save(input_path)
    
    if ext in allowed_images:
        metadata, score, risk_level = extract_metadata_image(input_path, filename)
        clean_image(input_path, output_path, share_safe)
    else:
        metadata, score, risk_level = extract_metadata_doc(input_path, filename, ext)
        clean_doc(input_path, output_path, ext)
        
    # User associated logic
    user_id = get_optional_user()
    if user_id:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('INSERT INTO history (id, user_id, filename, upload_date, score, risk_level) VALUES (?, ?, ?, ?, ?, ?)',
                  (file_id, user_id, file.filename, datetime.now(timezone.utc).isoformat(), score, risk_level))
        conn.commit()
        conn.close()
    
    return jsonify({
        "id": file_id,
        "filename": file.filename,
        "metadata": metadata,
        "score": score,
        "risk_level": risk_level,
        "clean_url": f"/api/download/{output_filename}"
    })

@app.route('/api/download/<filename>', methods=['GET'])
def download(filename):
    for f in os.listdir(CLEAN_FOLDER):
        if f == filename:
            return send_file(os.path.join(CLEAN_FOLDER, f), as_attachment=True, download_name=f"clean_{f}")
    return jsonify({"error": "File not found"}), 404

@app.route('/api/history', methods=['GET'])
@token_required
def get_history(user_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM history WHERE user_id = ? ORDER BY upload_date DESC', (user_id,))
    rows = c.fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

if __name__ == '__main__':
    app.run(debug=True, port=5000)
